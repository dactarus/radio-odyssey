import csv
import datetime
import re
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils import get_column_letter

SRC = "programmation/catalogue-titres-radioking.csv"
OUT = "programmation/catalogue-radio-odyssey.xlsx"

PURPLE = "7B2FBE"
DARK = "2D2D3A"
MUTED = "6C6C80"
FONT_NAME = "Calibri"

_ILLEGAL_RE = re.compile(r'[\x00-\x08\x0b\x0c\x0e-\x1f]')
def clean(v):
    if isinstance(v, str):
        return _ILLEGAL_RE.sub("'", v)
    return v

rows = []
with open(SRC, encoding="utf-8") as fh:
    reader = csv.DictReader(fh)
    for r in reader:
        rows.append((
            clean(r["Artist"].strip()),
            clean(r["Track"].strip()),
            int(r["Plays_total"] or 0),
            clean(r.get("Composer", "").strip()),
            int(r["Semestres_couverts"] or 0),
            clean(r.get("Variantes_fusionnees", "").strip()),
        ))

rows.sort(key=lambda x: -x[2])

def base_title(t):
    t2 = re.sub(r'\s*[\(\[][^)\]]*[\)\]]\s*', ' ', t)
    return re.sub(r'\s+', ' ', t2).strip().lower()

_groups_map = defaultdict(list)
for artist, track, plays, composer, sem, variantes in rows:
    _groups_map[(artist.strip().lower(), base_title(track))].append((artist, track, plays))
REVIEW_GROUPS = [items for items in _groups_map.values() if len(items) >= 2]
REVIEW_GROUPS.sort(key=lambda items: -sum(p for _, _, p in items))
N_A_TRANCHER = len(REVIEW_GROUPS)

by_artist = defaultdict(lambda: {"titres": 0, "passages": 0})
for artist, track, plays, composer, sem, variantes in rows:
    by_artist[artist]["titres"] += 1
    by_artist[artist]["passages"] += plays
artist_rows = sorted(by_artist.items(), key=lambda kv: -kv[1]["passages"])

n_fusions = sum(1 for r in rows if r[5])

wb = Workbook()

# ---------------------------------------------------------------- Résumé
ws0 = wb.active
ws0.title = "Résumé"
ws0.sheet_view.showGridLines = False

def set_title(ws, cell, text, size=16, color=PURPLE, bold=True):
    c = ws[cell]
    c.value = text
    c.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)

def label(ws, cell, text, bold=True, size=11, color=DARK):
    c = ws[cell]
    c.value = text
    c.font = Font(name=FONT_NAME, size=size, bold=bold, color=color)

set_title(ws0, "B2", "Catalogue des titres diffusés — Radio Odyssey")
ws0["B3"] = "Document interne — ne pas diffuser publiquement"
ws0["B3"].font = Font(name=FONT_NAME, size=10, italic=True, color=MUTED)

label(ws0, "B5", "Période couverte", size=11)
ws0["D5"] = "1er janvier 2025 → 2 septembre 2026 (continu, sans trou)"
ws0["D5"].font = Font(name=FONT_NAME, size=11, color=DARK)

label(ws0, "B6", "Dernière mise à jour", size=11)
ws0["D6"] = datetime.date(2026, 9, 3)
ws0["D6"].number_format = "dd/mm/yyyy"
ws0["D6"].font = Font(name=FONT_NAME, size=11, color=DARK)

label(ws0, "B7", "Source", size=11)
ws0["D7"] = "4 exports RadioKing \"fréquence\" (un par semestre), fusionnés — voir onglet Méthode"
ws0["D7"].font = Font(name=FONT_NAME, size=11, color=DARK)

stats = [
    ("Titres distincts", "=SUBTOTAL(103,Titres[Titre])"),
    ("Artistes / groupes distincts", "=SUBTOTAL(103,Artistes[Artiste])"),
    ("Total des passages observés", "=SUM(Titres[Passages])"),
    ("Titres diffusés sur les 4 semestres (en continu)", "=COUNTIF(Titres[Semestres couverts],4)"),
    ("Titres vus sur 1 seul semestre (ponctuels)", "=COUNTIF(Titres[Semestres couverts],1)"),
    ("Titres corrigés (crédits d'artiste fusionnés)", f"={n_fusions}"),
]
r = 9
label(ws0, f"B{r}", "Chiffres clés", size=13, color=PURPLE)
r += 1
for lbl, formula in stats:
    label(ws0, f"B{r}", lbl, bold=False)
    c = ws0[f"D{r}"]
    c.value = formula
    c.font = Font(name=FONT_NAME, size=11, bold=True, color=PURPLE)
    c.number_format = "#,##0"
    r += 1

r += 1
label(ws0, f"B{r}", "⚠️ Onglet \"À trancher\"", size=12, color=PURPLE)
r += 1
ws0[f"B{r}"] = (f"{N_A_TRANCHER} groupes de titres où un mot diffère (souvent un remix) restent "
                "à décider un par un : même enregistrement compté deux fois, ou deux versions "
                "réellement distinctes ? Impossible à trancher depuis les exports seuls — la "
                "décision revient au propriétaire, groupe par groupe (onglet \"À trancher\").")
ws0[f"B{r}"].font = Font(name=FONT_NAME, size=10.5, color=DARK)
ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
ws0.merge_cells(f"B{r}:H{r}")
ws0.row_dimensions[r].height = 30
r += 2
label(ws0, f"B{r}", "Comment lire ce fichier", size=13, color=PURPLE)
r += 1
notes = [
    "Onglet \"Titres\" : un titre par ligne — clique l'en-tête d'une colonne pour trier, ou utilise les flèches de filtre pour chercher un artiste précis.",
    "Onglet \"Artistes\" : agrégé par artiste — nombre de titres distincts et total de passages, recalculé automatiquement (formules) à partir de l'onglet Titres.",
    "\"Passages\" = nombre de fois où le titre a été diffusé sur la période couverte (les 4 semestres additionnés).",
    "\"Semestres couverts\" (1 à 4) = sur combien des 4 périodes de 6 mois le titre a été vu. 4 = diffusé en continu depuis janvier 2025 ; 1 = plus ponctuel.",
    "\"Variantes fusionnées\" : quand RadioKing a compté le même titre sous plusieurs crédits d'artiste différents (featuring déplacé, faute de frappe, casse différente...), les lignes ont été regroupées en une seule — cette colonne liste les crédits d'origine, pour que rien ne soit perdu. Vide = titre qui n'a pas eu besoin d'être corrigé.",
]
for n in notes:
    ws0[f"B{r}"] = "• " + n
    ws0[f"B{r}"].font = Font(name=FONT_NAME, size=10.5, color=DARK)
    ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.merge_cells(f"B{r}:H{r}")
    ws0.row_dimensions[r].height = 28
    r += 1

r += 1
label(ws0, f"B{r}", "Pour mettre à jour ce fichier", size=13, color=PURPLE)
r += 1
maj = [
    "Réexporter, depuis le manager RadioKing (Planification → Historique → Rapport des titres joués), un rapport \"fréquence\" pour la période écoulée depuis le 2 septembre 2026 — sans chevaucher les quatre exports déjà intégrés (voir onglet Méthode).",
    "Transmettre le nouveau fichier CSV pour que le catalogue et ce classeur soient refaits avec la période étendue.",
]
for n in maj:
    ws0[f"B{r}"] = "• " + n
    ws0[f"B{r}"].font = Font(name=FONT_NAME, size=10.5, color=DARK)
    ws0[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws0.merge_cells(f"B{r}:H{r}")
    ws0.row_dimensions[r].height = 28
    r += 1

ws0.column_dimensions["A"].width = 2
ws0.column_dimensions["B"].width = 20
ws0.column_dimensions["C"].width = 4
ws0.column_dimensions["D"].width = 60

# ---------------------------------------------------------------- Méthode
ws_m = wb.create_sheet("Méthode")
ws_m.sheet_view.showGridLines = False
set_title(ws_m, "B2", "Méthode de construction du catalogue", size=14)
r = 4
meth = [
    ("Quatre exports RadioKing", "\"fréquence\", un par semestre : 2025-01-01→2025-06-30, 2025-07-01→2025-12-31, 2026-01-01→2026-06-30, 2026-07-01→2026-09-02. Fenêtres continues et non chevauchantes, fournies par le propriétaire."),
    ("Fusion des semestres", "regroupement par (Artiste, Titre) exact ; \"Passages\" = somme du Play frequency des quatre exports pour ce couple — une simple addition suffit puisque les périodes ne se recoupent pas."),
    (f"Correction des crédits d'artiste ({n_fusions} titres)", "RadioKing traite parfois \"Artiste A\" et \"Artiste A, Artiste B\" comme deux entités distinctes pour un même enregistrement — le propriétaire ajuste en cours d'année en ne laissant que l'artiste principal dans la colonne Artiste et en déplaçant le featuring dans le titre (\"Ft ...\"). Ce qui fragmentait un même titre en plusieurs lignes a été identifié (chevauchement de crédit entre deux lignes de même titre) et regroupé, sur validation du propriétaire — voir la colonne \"Variantes fusionnées\" de l'onglet Titres pour la trace de chaque fusion. Les cas ambigus (mashups combinant deux chansons différentes, ex. \"Holiday\" x \"Don't Start Now\") ont volontairement été laissés distincts."),
    ("Ce que ce fichier n'est PAS", "un compteur RadioKing officiel en temps réel : c'est une reconstitution à partir d'exports, avec une correction manuelle documentée. Pour un chiffre certifié à un instant donné, un nouvel export direct depuis RadioKing reste la seule source qui fasse foi."),
    ("Différence avec artists.js (le site)", "le site radio-odyssey.com n'affiche que 147 fiches artistes, au plus 8 titres chacune, sur une fenêtre glissante de 3 mois (14 mai → 13 août 2026) — pour rester comparable dans le temps sur les fiches publiées. Ce classeur-ci n'a pas cette contrainte : tout artiste et tout titre vus dans les quatre exports, sans plafond, sur 20 mois."),
    ("Confidentialité", "ce fichier n'est ni publié sur le site, ni suivi dans le dépôt Git du projet (exclu via .gitignore) — à garder en local, hors accès public : c'est la programmation réelle et complète de l'antenne."),
]
for titre, txt in meth:
    ws_m[f"B{r}"] = titre
    ws_m[f"B{r}"].font = Font(name=FONT_NAME, size=11, bold=True, color=PURPLE)
    r += 1
    ws_m[f"B{r}"] = txt
    ws_m[f"B{r}"].font = Font(name=FONT_NAME, size=10.5, color=DARK)
    ws_m[f"B{r}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_m.merge_cells(f"B{r}:H{r}")
    ws_m.row_dimensions[r].height = 56
    r += 2
ws_m.column_dimensions["A"].width = 2
ws_m.column_dimensions["B"].width = 100

# ---------------------------------------------------------------- Titres
ws1 = wb.create_sheet("Titres")
headers1 = ["Artiste", "Titre", "Passages", "Semestres couverts", "Compositeur", "Variantes fusionnées"]
ws1.append(headers1)
for artist, track, plays, composer, sem, variantes in rows:
    ws1.append([artist, track, plays, sem, composer, variantes])

n1 = len(rows) + 1
widths1 = [30, 46, 12, 18, 22, 50]
for i, w in enumerate(widths1, start=1):
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.freeze_panes = "A2"

tab1 = Table(displayName="Titres", ref=f"A1:F{n1}")
tab1.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
ws1.add_table(tab1)

# ---------------------------------------------------------------- Artistes
ws2 = wb.create_sheet("Artistes")
headers2 = ["Artiste", "Titres distincts", "Total des passages"]
ws2.append(headers2)
for artist, agg in artist_rows:
    ws2.append([artist, None, None])

n2 = len(artist_rows) + 1
for row_idx in range(2, n2 + 1):
    ws2[f"B{row_idx}"] = f"=COUNTIF(Titres[Artiste],A{row_idx})"
    ws2[f"C{row_idx}"] = f"=SUMIF(Titres[Artiste],A{row_idx},Titres[Passages])"

widths2 = [34, 16, 18]
for i, w in enumerate(widths2, start=1):
    ws2.column_dimensions[get_column_letter(i)].width = w
ws2.freeze_panes = "A2"

tab2 = Table(displayName="Artistes", ref=f"A1:C{n2}")
tab2.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False)
ws2.add_table(tab2)

# Entêtes colorées (au-dessus du style de table, garanti visible)
fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
for sheet in (ws1, ws2):
    for cell in sheet[1]:
        cell.fill = fill
        cell.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
        cell.alignment = Alignment(vertical="center")
    sheet.row_dimensions[1].height = 20

# ---------------------------------------------------------------- À trancher
from openpyxl.styles import Border, Side

ws3 = wb.create_sheet("À trancher")
ws3.sheet_view.showGridLines = False

ws3["B2"] = "Titres à trancher : fusionner ou garder séparés ?"
ws3["B2"].font = Font(name=FONT_NAME, size=14, bold=True, color=PURPLE)
ws3["B3"] = ("Même artiste, même titre de base, mais un mot diffère (souvent un nom de remix) — "
             "impossible de savoir depuis les exports seuls si c'est le même enregistrement compté "
             "deux fois ou deux versions réellement distinctes dans la playlist.")
ws3["B3"].font = Font(name=FONT_NAME, size=10.5, color=DARK, italic=True)
ws3["B3"].alignment = Alignment(wrap_text=True, vertical="top")
ws3.merge_cells("B3:F3")
ws3.row_dimensions[3].height = 32
ws3["B4"] = "Pour chaque groupe : écris \"Fusionner\" ou \"Garder séparé\" dans la colonne Décision (une seule fois par groupe, sur sa première ligne)."
ws3["B4"].font = Font(name=FONT_NAME, size=10.5, color=DARK, bold=True)
ws3.merge_cells("B4:F4")
ws3.row_dimensions[4].height = 18

headers3 = ["N°", "Artiste", "Titre (variante)", "Passages", "Total du groupe", "Décision (Fusionner / Garder séparé)"]
header_row = 6
for ci, h in enumerate(headers3, start=2):
    c = ws3.cell(row=header_row, column=ci, value=h)
    c.font = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
    c.fill = PatternFill(start_color=PURPLE, end_color=PURPLE, fill_type="solid")
    c.alignment = Alignment(vertical="center", wrap_text=True)
ws3.row_dimensions[header_row].height = 32

thin = Side(style="thin", color="D9D2E9")
cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)
yellow_fill = PatternFill(start_color="FFF3B0", end_color="FFF3B0", fill_type="solid")

rr = header_row + 1
for gid, items in enumerate(REVIEW_GROUPS, start=1):
    start_row = rr
    total = sum(p for _, _, p in items)
    for artist, track, plays in items:
        ws3.cell(row=rr, column=3, value=artist).font = Font(name=FONT_NAME, size=10.5, color=DARK)
        ws3.cell(row=rr, column=4, value=track).font = Font(name=FONT_NAME, size=10.5, color=DARK)
        ws3.cell(row=rr, column=5, value=plays).font = Font(name=FONT_NAME, size=10.5, color=DARK)
        for col in (2, 3, 4, 5, 6, 7):
            ws3.cell(row=rr, column=col).border = cell_border
        rr += 1
    end_row = rr - 1
    ws3.cell(row=start_row, column=2, value=gid).font = Font(name=FONT_NAME, size=10.5, bold=True, color=PURPLE)
    ws3.cell(row=start_row, column=6, value=total).font = Font(name=FONT_NAME, size=10.5, bold=True, color=PURPLE)
    dcell = ws3.cell(row=start_row, column=7)
    dcell.fill = yellow_fill
    dcell.font = Font(name=FONT_NAME, size=10.5, color=DARK)
    dcell.alignment = Alignment(vertical="center")
    if end_row > start_row:
        ws3.merge_cells(start_row=start_row, start_column=2, end_row=end_row, end_column=2)
        ws3.merge_cells(start_row=start_row, start_column=6, end_row=end_row, end_column=6)
        ws3.merge_cells(start_row=start_row, start_column=7, end_row=end_row, end_column=7)
        ws3.cell(row=start_row, column=2).alignment = Alignment(vertical="center", horizontal="center")
        ws3.cell(row=start_row, column=6).alignment = Alignment(vertical="center", horizontal="center")

widths3 = {2: 6, 3: 24, 4: 52, 5: 12, 6: 14, 7: 34}
for col, w in widths3.items():
    ws3.column_dimensions[get_column_letter(col)].width = w
ws3.freeze_panes = f"C{header_row+1}"

order = [ws0, ws1, ws3, ws2, ws_m]
wb._sheets = order
wb.active = 0

wb.save(OUT)
print("Écrit:", OUT, "-", len(rows), "titres,", len(artist_rows), "artistes,", n_fusions, "corrigés,", N_A_TRANCHER, "groupes à trancher")
